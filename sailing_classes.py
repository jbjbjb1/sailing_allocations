# Classes for main.py

import random
import os
import copy

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter


class Camp:
    def __init__(self, df_camper, df_leader, df_boat, df_schedule):
        """ Loads input information needed to build entire program for camp. """
        self.df_camper = df_camper
        self.df_leader = df_leader
        self.df_boat = df_boat
        self.df_schedule = df_schedule
        self.duty_groups = self.df_camper["Duty Group"].unique().tolist()
        # Write functions to be executed when instance of class is created      


        def availability_count():
            """ Calculates availability of leaders, boats and campers. """   
            leader_avail = self.df_leader[(self.df_leader["Available"] != "n") 
                                        & ( (self.df_leader["SkippersTicket"] == "y") 
                                        | (self.df_leader["CompVag"] == "y")
                                        | (self.df_leader["CompCat"] == "y") 
                                        | (self.df_leader["CompRaceControl"] == "y")
                                        )]
            craft_avail = self.df_boat[self.df_boat["Available"] != "n"]
            camper_avail = self.df_camper[self.df_camper["Available"] != "n"]
            return leader_avail, craft_avail, camper_avail

        
        def initiate_balance_log():
            """ Initiate the balance log with the leaders, campers and boats for this allocation. """
            balance_log_columns = self.df_boat["Type"].unique().tolist()
            list_of_leaders_campers = self.leader_avail.index.tolist() + self.camper_avail.index.tolist()
            balance_log = pd.DataFrame(0, index=list_of_leaders_campers, columns=balance_log_columns)
            balance_log.index.names = ["Names"]
            return balance_log
        
        
        def generate_allocations():
            """ Iterate through timeslots in df_schedule and give an allocations. """
            allocations = []
            for timeslot in list(df_schedule.index.values):
                new_allocation = Allocations(self, timeslot)      # create new instance of allocation class               
                allocations.append(new_allocation)
            return allocations, self.balance_log
        
        
        # Call functions to run when instance of class is created
        self.leader_avail, self.craft_avail, self.camper_avail = availability_count()
        self.balance_log = initiate_balance_log()
        self.allocations, self.balance_log = generate_allocations()
    

    def numbers(self):
        """ Prints numbers of people on camp. """
        print('Leaders available:', len(self.leader_avail))
        print('Craft count available (inc. beach):', len(self.craft_avail))
        print('Camper positions available (inc. beach):', self.craft_avail["CamperCapacity"].sum())
        print('Campers available:', len(self.camper_avail))


    def export_plan(self):
        """ Exports entire generated camp plan to Excel Workbook. """
        wb = Workbook()                 # create new Excel workbook
        ws = []
        ws.append(wb.active)            # gets the current Excel worksheet
        ws[-1].title = 'Overview'       # [-1] selects the latest sheet
        ws[-1]['A1'] = 'This file contains allocations for the camp.'   
        for allocation in self.allocations:
            allocation_name = allocation.day + ' ' + allocation.time
            ws.append(wb.create_sheet(allocation_name))     # create a new sheet with allocation name
            # Set column width
            column_width = 14
            # Create border style
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
            # Create formatting at top of A4 page
            ws[-1]['A1'] = "SAILING ALLOCATION"
            ws[-1].column_dimensions["A"].width = column_width
            ws[-1]['H1'] = "DATE:"
            ws[-1]['I1'] = allocation_name
            ws[-1]['A3'] = "Boat"
            ws[-1]['A3'].border = thin_border
            ws[-1]['A4'] = "Skipper"
            ws[-1]['A4'].border = thin_border
            # Create row for Boat
            for row in [3]:
                for col in range(2, len(allocation.crew)+2):
                    a = ws[-1].cell(column=col, row=row, value=allocation.crew[col-2].craft)    # put boat name
                    a.border = thin_border                                                      # apply border
                    # Apply fixed column widths, make it wider for Beach
                    if allocation.crew[col-2].craft == "Beach":
                        ws[-1].column_dimensions[get_column_letter(col)].width = column_width + 18       
                    else:
                        ws[-1].column_dimensions[get_column_letter(col)].width = column_width   
            # Create row for Skipper
            for row in [4]:
                for col in range(2, len(allocation.crew)+2):
                    b = ws[-1].cell(column=col, row=row, value=allocation.crew[col-2].leader)    # put boat name
                    b.border = thin_border                                                       # apply border
            # Create rows for Sessions
            for session in allocation.sessions:
                _ = ws[-1].cell(column=1, row=session.session_no+4, value="Session 0{}".format(session.session_no))     # put session name
                _.border = thin_border                                                                                  # apply border
                # Add campers to cells
                for row in [session.session_no+4]:
                    for col in range(2, len(allocation.crew)+2):
                        # For beach only, put 3 campers on each line
                        if session.crew[col-2].craft == "Beach":
                            campers = session.crew[col-2].campers
                            campers_string = ""
                            # Loop over groups of 3
                            for i in range(int(len(campers)/3)):
                                campers_string += ', '.join(campers[i:i+3]) + '\n'      # loop over groups of 3
                            # Remove trailing \n if number of campers divisible by 3
                            if len(campers)/3 == 0:
                                campers_string = campers_string[:-2]                    
                            # If not divisible by 3, add remainder
                            else:
                                campers_string += ', '.join(campers[int(len(campers)/3)*3:])                 # if there are extra, add remaining
                            c = ws[-1].cell(column=col, row=row, value=campers_string)
                        # Otherwise put a single camper on each line
                        else:
                            c = ws[-1].cell(column=col, row=row, value='\n'.join(session.crew[col-2].campers))
                        c.border = thin_border                                                                          # apply border
                        c.alignment = Alignment(wrap_text = True)                                                       # wrap text for campers
        # Save the file to Excel
        saved = False
        while saved is False:
            try:
                # Save the file
                filename = "sailing_allocation_output.xlsx"
                wb.save(filename)
                # Print status to terminal and open file
                print('Export sailing_allocation_output.xlsx complete.')
                saved = True
            except PermissionError:
                input('Please close sailing_allocation_output.xlsx and press enter to continue.')


    def export_balance_log(self):
        """ Export to Excel the turns leaders and campers had on different craft types. """
        saved = False
        while saved is False:
            try:
                # Save the file
                with pd.ExcelWriter('sailing_allocation_output_turns.xlsx') as writer:
                    balance_log_leaders = self.balance_log[self.balance_log.index.isin(self.leader_avail.index.tolist())]
                    balance_log_leaders.to_excel(writer, sheet_name='Leaders_turns')   # leaders in balance_log
                    balance_log_campers = self.balance_log[self.balance_log.index.isin(self.camper_avail.index.tolist())]
                    balance_log_campers.to_excel(writer, sheet_name='Campers_turns')   # campers in balance_log
                print('Export sailing_allocation_output_turns.xlsx complete.')
                saved = True
            except PermissionError:
                input('Please close sailing_allocation_output_turns.xlsx and press enter to continue.')
        

class Crew:
    def __init__(self, leader, craft, craft_type, capacity):
        """ Holds the leader, craft type and campers involved for a specific session. """
        self.leader = leader
        self.craft = craft
        self.craft_type = craft_type
        self.capacity = capacity
        self.campers = []
    

    def __str__(self):
        return "Crew leader: " + self.leader + ", Craft: " + self.craft + ", Craft type: " + self.craft_type + ", Capacity: " + str(self.capacity) + ", Campers: " + ', '.join(self.campers)


class Allocations:
    def __init__(self, camp, timeslot):
        """ Creates the program for the allocation based on the duty groups assigned (blank means everyone). """
        self.camp = camp
        self.day = timeslot.split()[0]
        self.time = timeslot.split()[1]
        # If duty group is blank assign all duty groups, otherwise change string of duty groups to list.
        self.duty_groups = camp.df_schedule.loc[timeslot, "Duty Group"]
        if self.duty_groups is np.nan:
            self.duty_groups = camp.duty_groups
        else:
            self.duty_groups = self.duty_groups.split(',')
            self.duty_groups = [group.strip() for group in self.duty_groups] 

        # Write functions to be executed when instance of class is created           
        
        def update_balance_log(session_log):
            """ Takes the list of assigned leaders and campers from a session and updates the balance_log. """
            for entry in session_log:
                for name, craft_type in entry.items():
                    # Update the balance log for what each leader and camper was assigned to 
                    self.balance_log.loc[name, craft_type] += 1
            return self.balance_log
            

        def assign_craft(craft_type, craft_avail):
            """ Assign the next craft in a given craft type. """
            craft = craft_avail.index[craft_avail["Type"] == craft_type][0]    # select first craft on list
            craft_avail = craft_avail.drop([craft])     # drop selected craft
            return craft, craft_avail


        def leader_balance(craft_type, leader_avail):
            """ Use balance to assign leader to a craft type. """
            # Below are masks to select appropriate leader from list
            leader_requirement = {"Rescue Boat": r'(leader_avail["SkippersTicket"] == "y")',
                                    "Race Control": r'(leader_avail["CompRaceControl"] == "y")',
                                    "Vagabond": r'(leader_avail["CompVag"] == "y")',
                                    "Cat": r'(leader_avail["CompCat"] == "y")',
                                    "Beach": '(leader_avail["Available"] != "n")'}   # dummy requirement for beach, as person already available
            # Get the list of leaders that meet the requirements for this craft
            leader_avail_req = leader_avail[eval(leader_requirement[craft_type])]
            balance_log_leader_avail_req = leader_avail_req.merge(camp.balance_log[craft_type], how='left', left_index=True, right_index=True)   # merge number of turns onto camper_avail df
            least_turns = balance_log_leader_avail_req[craft_type].min()     # find the least turns someone has had in that craft_type
            least_turns_leaders = balance_log_leader_avail_req.index[(balance_log_leader_avail_req[craft_type] == least_turns)].tolist()    # get a list of all people who have had that many turns
            leader = random.choice(least_turns_leaders)   # randomly select someone from that list
            leader_avail = leader_avail.drop([leader])  # drop selected leader
            return leader, leader_avail


        def initiate_crew():
            """ Assign crew for allocation (i.e. leader/boat allocation). """   
            # The following variable defines the order & minimum number of craft required
            allocate_define = {"Rescue Boat": 2,
                                "Race Control": 1,
                                "Vagabond": 99,
                                "Cat": 99,
                                "Beach": 99}
            # Assign local values for function to craft and leader availability
            craft_avail = self.craft_avail
            leader_avail = self.leader_avail
            crew = []
            for craft_type, num in allocate_define.items():
                while num > 0:
                    try:
                        craft, craft_avail = assign_craft(craft_type, craft_avail)
                        leader, leader_avail = leader_balance(craft_type, leader_avail)
                    except IndexError:
                        # If there are no craft or leaders available, Pandas will throw an IndexError in the respective function
                        #print('--------------------IndexError--------------------')
                        break
                    capacity = self.camp.df_boat.loc[craft, "CamperCapacity"]
                    crew.append(Crew(leader, craft, craft_type, capacity))
                    num += -1
            return crew
        

        def generate_session():
            """ Iterate through three sessions for campers/leaders/boats. """
            sessions = []
            for session_no in range(1, 4):
                # copy.deepcopy() is needed otherwise all new self.crew will be linked and overwrite each other
                new_session = Session(self, session_no, copy.deepcopy(self.crew), self.balance_log)           # create new instance of allocation class
                self.balance_log = update_balance_log(new_session.session_log)                                # add the allocations of that session to the balance log               
                sessions.append(new_session)
            return sessions, self.balance_log

        # Call functions to run when instance of class is created        
        self.leader_avail, self.craft_avail, self.camper_avail = self.camp.leader_avail, self.camp.craft_avail, self.camp.camper_avail
        self.balance_log = self.camp.balance_log                    # reference the balance log from the allocation class
        self.crew = initiate_crew()
        self.sessions, self.balance_log = generate_session()    # save sessions created and updated balance log   


class Session:
    def __init__(self, allocation, session_no, crew, balance_log):
        """ Fills a single session for the list of duty groups. """
        self.allocation = allocation
        self.session_no = session_no
        self.crew = crew
        self.session_log = [{crew.leader: crew.craft_type} for crew in self.crew] # append the leaders for the session to the log
        # Write functions to be executed when instance of class is created      
        

        def camper_balance(craft_type, camper_avail):
            """ Use balance to assign camper to a craft type. """
            balance_log_camper_avail = camper_avail.merge(balance_log[craft_type], how='left', left_index=True, right_index=True)   # merge number of turns onto camper_avail df
            least_turns = balance_log_camper_avail[craft_type].min()     # find the least turns someone has had in that craft_type
            least_turns_campers = balance_log_camper_avail.index[(balance_log_camper_avail[craft_type] == least_turns)].tolist()  # get a list of all people who have had that many turns
            camper = random.choice(least_turns_campers)     # randomly select someone from that list
            camper_avail = camper_avail.drop([camper])      # drop selected camper
            return camper, camper_avail

        
        def initiate_crew():
            """ Assign crew for session (i.e. camper). """   
            camper_avail = allocation.camper_avail
            for crew in self.crew:
                campers = []
                capacity = crew.capacity   # repeat as many times as possible
                while capacity > 0:
                    try:
                        camper, camper_avail = camper_balance(crew.craft_type, camper_avail)
                        self.session_log.append({camper: crew.craft_type})   # save a record of who was assigned to what craft type
                    except IndexError:
                        # If there are no craft or leaders available, Pandas will throw an IndexError in the respective function
                        # print('--------------------IndexError--------------------')
                        break
                    campers.append(camper)
                    capacity += -1
                crew.campers = campers
            return self.crew, self.session_log
            

        # Call functions to run when instance of class is created
        self.crew, self.session_log = initiate_crew()     # overwrites crew with campers added and session log of who was assigned where